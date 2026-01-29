from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from extensions import db
from models import Holding
from datetime import datetime
import csv
import io

portfolio_bp = Blueprint('portfolio', __name__, url_prefix='/portfolio')

@portfolio_bp.route('/')
@login_required
def index():
    holdings = Holding.query.filter_by(user_id=current_user.id).all()
    
    # Consolidate holdings by symbol (combine multiple purchases of same stock)
    consolidated = {}
    for h in holdings:
        symbol = h.symbol
        if symbol not in consolidated:
            consolidated[symbol] = {
                'ids': [h.id],
                'symbol': symbol,
                'company_name': h.company_name,
                'quantity': h.quantity,
                'total_cost': h.quantity * h.buy_price,
                'sector': h.sector,
                'earliest_buy_date': h.buy_date,
            }
        else:
            consolidated[symbol]['ids'].append(h.id)
            consolidated[symbol]['quantity'] += h.quantity
            consolidated[symbol]['total_cost'] += h.quantity * h.buy_price
            # Keep the earliest buy date
            if h.buy_date < consolidated[symbol]['earliest_buy_date']:
                consolidated[symbol]['earliest_buy_date'] = h.buy_date
            # Update company name if empty
            if not consolidated[symbol]['company_name'] and h.company_name:
                consolidated[symbol]['company_name'] = h.company_name
            # Update sector if empty
            if not consolidated[symbol]['sector'] and h.sector:
                consolidated[symbol]['sector'] = h.sector
    
    # Calculate weighted average buy price for each consolidated holding
    for symbol, data in consolidated.items():
        data['avg_buy_price'] = data['total_cost'] / data['quantity'] if data['quantity'] > 0 else 0
    
    # Calculate portfolio stats using consolidated data
    total_invested = sum(data['total_cost'] for data in consolidated.values())
    
    # Fetch current prices, dividend info, and day changes
    from services.stock_service import get_current_prices, get_dividend_info, get_day_changes
    symbols = list(consolidated.keys())
    current_prices = get_current_prices(symbols)
    dividend_info = get_dividend_info(symbols)
    day_changes = get_day_changes(symbols)
    
    holdings_data = []
    total_current = 0
    total_dividend = 0
    total_day_change = 0
    
    for symbol, data in consolidated.items():
        # Get price data with currency info
        price_data = current_prices.get(symbol, {
            'price': data['avg_buy_price'],
            'currency': '₹',
            'currency_code': 'INR',
            'market': 'IN'
        })
        
        current_price = price_data['price'] if isinstance(price_data, dict) else price_data
        currency = price_data.get('currency', '₹') if isinstance(price_data, dict) else '₹'
        
        current_value = data['quantity'] * current_price
        gain_loss = current_value - data['total_cost']
        gain_loss_pct = ((current_price - data['avg_buy_price']) / data['avg_buy_price']) * 100 if data['avg_buy_price'] > 0 else 0
        
        # Calculate dividend earned (quantity × annual dividend per share)
        dividend_per_share = dividend_info.get(symbol, 0)
        dividend_earned = data['quantity'] * dividend_per_share
        total_dividend += dividend_earned
        
        # Get day change for this stock
        day_change_data = day_changes.get(symbol, {'change': 0, 'change_pct': 0})
        day_change = day_change_data['change'] * data['quantity']
        day_change_pct = day_change_data['change_pct']
        total_day_change += day_change
        
        total_current += current_value
        holdings_data.append({
            'ids': data['ids'],  # List of all holding IDs for this symbol
            'id': data['ids'][0],  # Primary ID (first holding)
            'symbol': symbol,
            'company_name': data['company_name'],
            'quantity': data['quantity'],
            'buy_price': data['avg_buy_price'],  # Weighted average price
            'total_invested': data['total_cost'],
            'current_price': current_price,
            'current_value': current_value,
            'gain_loss': gain_loss,
            'gain_loss_pct': gain_loss_pct,
            'sector': data['sector'],
            'buy_date': data['earliest_buy_date'],  # Earliest buy date
            'currency': currency,
            'market': price_data.get('market', 'IN') if isinstance(price_data, dict) else 'IN',
            'dividend_per_share': dividend_per_share,
            'dividend_earned': dividend_earned,
            'lot_count': len(data['ids']),  # Number of separate lots
            'day_change': day_change,
            'day_change_pct': day_change_pct
        })
    
    # Sector distribution
    sector_data = {}
    for h in holdings_data:
        sector = h['sector'] or 'Other'
        sector_data[sector] = sector_data.get(sector, 0) + h['current_value']
    
    # Calculate day change percentage for portfolio
    total_day_change_pct = (total_day_change / (total_current - total_day_change)) * 100 if (total_current - total_day_change) > 0 else 0
    
    return render_template('portfolio/index.html', 
                         holdings=holdings_data,
                         total_invested=total_invested,
                         total_current=total_current,
                         total_gain_loss=total_current - total_invested,
                         total_dividend=total_dividend,
                         total_day_change=total_day_change,
                         total_day_change_pct=total_day_change_pct,
                         sector_data=sector_data)

@portfolio_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_stock():
    if request.method == 'POST':
        from services.stock_service import validate_indian_stock
        
        symbol = request.form.get('symbol', '').upper().strip()
        company_name = request.form.get('company_name', '').strip()
        quantity = float(request.form.get('quantity', 0))
        buy_price = float(request.form.get('buy_price', 0))
        buy_date = datetime.strptime(request.form.get('buy_date'), '%Y-%m-%d').date()
        sector = request.form.get('sector', '').strip()
        
        # Validate Indian stock
        if not validate_indian_stock(symbol):
            flash('Only Indian stocks are supported. Please use NSE (.NS) or BSE (.BO) symbols (e.g., TCS.NS, RELIANCE.NS)', 'error')
            return redirect(url_for('portfolio.add_stock'))
        
        if not symbol or quantity <= 0 or buy_price <= 0:
            flash('Please fill in all required fields correctly', 'error')
            return redirect(url_for('portfolio.add_stock'))
        
        holding = Holding(
            user_id=current_user.id,
            symbol=symbol,
            company_name=company_name,
            quantity=quantity,
            buy_price=buy_price,
            buy_date=buy_date,
            sector=sector
        )
        
        db.session.add(holding)
        db.session.commit()
        
        flash(f'{symbol} added to your portfolio!', 'success')
        return redirect(url_for('portfolio.index'))
    
    return render_template('portfolio/add_stock.html')

@portfolio_bp.route('/batch', methods=['GET', 'POST'])
@login_required
def batch_upload():
    if request.method == 'POST':
        stocks_data = request.form.get('stocks_data', '')
        
        lines = stocks_data.strip().split('\n')
        added = 0
        
        for line in lines:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) >= 4:
                try:
                    holding = Holding(
                        user_id=current_user.id,
                        symbol=parts[0].upper(),
                        company_name=parts[1] if len(parts) > 4 else '',
                        quantity=float(parts[2] if len(parts) > 4 else parts[1]),
                        buy_price=float(parts[3] if len(parts) > 4 else parts[2]),
                        buy_date=datetime.strptime(parts[4] if len(parts) > 4 else parts[3], '%Y-%m-%d').date(),
                        sector=parts[5] if len(parts) > 5 else ''
                    )
                    db.session.add(holding)
                    added += 1
                except (ValueError, IndexError):
                    continue
        
        db.session.commit()
        flash(f'Successfully added {added} stocks to your portfolio!', 'success')
        return redirect(url_for('portfolio.index'))
    
    return render_template('portfolio/batch_upload.html')

@portfolio_bp.route('/upload-csv', methods=['POST'])
@login_required
def upload_csv():
    if 'csv_file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('portfolio.batch_upload'))
    
    file = request.files['csv_file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('portfolio.batch_upload'))
    
    if not file.filename.endswith('.csv'):
        flash('Please upload a CSV file', 'error')
        return redirect(url_for('portfolio.batch_upload'))
    
    try:
        content = file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(content))
        
        added = 0
        for row in csv_reader:
            try:
                holding = Holding(
                    user_id=current_user.id,
                    symbol=row.get('symbol', row.get('Symbol', '')).upper().strip(),
                    company_name=row.get('company_name', row.get('Company', '')),
                    quantity=float(row.get('quantity', row.get('Quantity', 0))),
                    buy_price=float(row.get('buy_price', row.get('Price', row.get('Buy Price', 0)))),
                    buy_date=datetime.strptime(
                        row.get('buy_date', row.get('Date', row.get('Buy Date', datetime.now().strftime('%Y-%m-%d')))),
                        '%Y-%m-%d'
                    ).date(),
                    sector=row.get('sector', row.get('Sector', ''))
                )
                db.session.add(holding)
                added += 1
            except (ValueError, KeyError):
                continue
        
        db.session.commit()
        flash(f'Successfully imported {added} stocks from CSV!', 'success')
    except Exception as e:
        flash(f'Error processing CSV: {str(e)}', 'error')
    
    return redirect(url_for('portfolio.index'))

@portfolio_bp.route('/api/holdings/<int:id>', methods=['PUT'])
@login_required
def update_holding(id):
    """Update a holding via AJAX"""
    try:
        holding = Holding.query.get_or_404(id)
        
        # Verify ownership
        if holding.user_id != current_user.id:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        
        # Update fields
        if 'quantity' in data:
            quantity = float(data['quantity'])
            if quantity <= 0:
                return jsonify({'success': False, 'error': 'Quantity must be greater than 0'}), 400
            holding.quantity = quantity
        
        if 'buy_price' in data:
            buy_price = float(data['buy_price'])
            if buy_price <= 0:
                return jsonify({'success': False, 'error': 'Buy price must be greater than 0'}), 400
            holding.buy_price = buy_price
        
        if 'buy_date' in data:
            holding.buy_date = datetime.strptime(data['buy_date'], '%Y-%m-%d').date()
        
        if 'sector' in data:
            holding.sector = data['sector']
        
        db.session.commit()
        
        # Get updated current price
        from services.stock_service import get_current_prices
        current_prices = get_current_prices([holding.symbol])
        price_data = current_prices.get(holding.symbol, {
            'price': holding.buy_price,
            'currency': '₹',
            'currency_code': 'INR',
            'market': 'IN'
        })
        
        current_price = price_data['price']
        current_value = holding.quantity * current_price
        gain_loss = current_value - (holding.quantity * holding.buy_price)
        gain_loss_pct = (gain_loss / (holding.quantity * holding.buy_price)) * 100 if holding.buy_price else 0
        
        return jsonify({
            'success': True,
            'message': f'{holding.symbol} updated successfully',
            'holding': {
                'id': holding.id,
                'symbol': holding.symbol,
                'company_name': holding.company_name,
                'quantity': holding.quantity,
                'buy_price': holding.buy_price,
                'buy_date': holding.buy_date.strftime('%Y-%m-%d'),
                'sector': holding.sector,
                'current_price': current_price,
                'current_value': current_value,
                'gain_loss': gain_loss,
                'gain_loss_pct': gain_loss_pct
            }
        })
        
    except ValueError as e:
        return jsonify({'success': False, 'error': 'Invalid data format'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@portfolio_bp.route('/api/holdings/<int:id>', methods=['DELETE'])
@login_required
def delete_holding_ajax(id):
    """Delete a holding via AJAX"""
    try:
        holding = Holding.query.get_or_404(id)
        
        # Verify ownership
        if holding.user_id != current_user.id:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
        symbol = holding.symbol
        db.session.delete(holding)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{symbol} removed from your portfolio'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# Keep the old delete route for backward compatibility
@portfolio_bp.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_stock(id):
    holding = Holding.query.get_or_404(id)
    
    if holding.user_id != current_user.id:
        flash('Unauthorized', 'error')
        return redirect(url_for('portfolio.index'))
    
    symbol = holding.symbol
    db.session.delete(holding)
    db.session.commit()
    
    flash(f'{symbol} removed from your portfolio', 'success')
    return redirect(url_for('portfolio.index'))


@portfolio_bp.route('/api/corporate-actions/<symbol>')
@login_required
def get_corporate_actions_api(symbol):
    """Get corporate actions for a stock (dividends, splits, bonus, etc.)"""
    try:
        from services.stock_service import get_corporate_actions
        
        data = get_corporate_actions(symbol)
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@portfolio_bp.route('/api/dividend-calendar')
@login_required
def get_dividend_calendar_api():
    """Get dividend calendar for user's holdings"""
    try:
        from services.stock_service import get_dividend_calendar
        
        # Get user's holdings
        holdings = Holding.query.filter_by(user_id=current_user.id).all()
        
        if not holdings:
            return jsonify({
                'success': True,
                'data': [],
                'message': 'No holdings found'
            })
        
        # Consolidate by symbol
        consolidated = {}
        for h in holdings:
            if h.symbol not in consolidated:
                consolidated[h.symbol] = {
                    'symbol': h.symbol,
                    'quantity': h.quantity,
                    'company_name': h.company_name
                }
            else:
                consolidated[h.symbol]['quantity'] += h.quantity
        
        holdings_data = list(consolidated.values())
        dividend_events = get_dividend_calendar(holdings_data)
        
        return jsonify({
            'success': True,
            'data': dividend_events
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@portfolio_bp.route('/api/performance-history')
@login_required
def get_performance_history_api():
    """Get portfolio performance history for charting"""
    try:
        from services.stock_service import get_portfolio_history
        
        period = request.args.get('period', '1M')
        
        # Get user's holdings
        holdings = Holding.query.filter_by(user_id=current_user.id).all()
        
        if not holdings:
            return jsonify({
                'success': True,
                'data': {'dates': [], 'values': [], 'change': 0, 'change_pct': 0},
                'message': 'No holdings found'
            })
        
        holdings_data = [
            {'symbol': h.symbol, 'quantity': h.quantity, 'buy_date': h.buy_date}
            for h in holdings
        ]
        
        history = get_portfolio_history(holdings_data, period)
        
        return jsonify({
            'success': True,
            'data': history
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@portfolio_bp.route('/api/ai-review')
@login_required
def get_ai_review_api():
    """Get AI-powered portfolio review using Gemini"""
    try:
        from services.ai_service import get_portfolio_review_gemini, get_stock_recommendations_gemini
        from services.stock_service import get_current_prices, get_dividend_info
        
        # Get user's holdings
        holdings = Holding.query.filter_by(user_id=current_user.id).all()
        
        if not holdings:
            return jsonify({
                'success': False,
                'error': 'No holdings found. Add stocks to your portfolio first.'
            })
        
        # Consolidate holdings by symbol
        consolidated = {}
        for h in holdings:
            symbol = h.symbol
            if symbol not in consolidated:
                consolidated[symbol] = {
                    'symbol': symbol,
                    'company_name': h.company_name,
                    'quantity': h.quantity,
                    'total_cost': h.quantity * h.buy_price,
                    'sector': h.sector,
                }
            else:
                consolidated[symbol]['quantity'] += h.quantity
                consolidated[symbol]['total_cost'] += h.quantity * h.buy_price
        
        # Fetch current prices
        symbols = list(consolidated.keys())
        current_prices = get_current_prices(symbols)
        
        # Build holdings data
        holdings_data = []
        total_invested = 0
        total_current = 0
        sector_data = {}
        
        for symbol, data in consolidated.items():
            price_data = current_prices.get(symbol, {'price': 0})
            current_price = price_data['price'] if isinstance(price_data, dict) else price_data
            current_value = data['quantity'] * current_price
            gain_loss = current_value - data['total_cost']
            
            total_invested += data['total_cost']
            total_current += current_value
            
            sector = data['sector'] or 'Other'
            sector_data[sector] = sector_data.get(sector, 0) + current_value
            
            holdings_data.append({
                'symbol': symbol,
                'company_name': data['company_name'],
                'quantity': data['quantity'],
                'total_invested': data['total_cost'],
                'current_value': current_value,
                'current_price': current_price,
                'gain_loss': gain_loss,
                'sector': sector
            })
        
        portfolio_stats = {
            'total_invested': total_invested,
            'total_current': total_current,
            'total_gain_loss': total_current - total_invested,
            'sector_data': sector_data
        }
        
        # Get AI portfolio-level review
        portfolio_review = get_portfolio_review_gemini(holdings_data, portfolio_stats)
        
        # Get AI stock-level recommendations with buy/sell/hold advice
        stock_recommendations = get_stock_recommendations_gemini(holdings_data)
        
        # Combine results
        result = {
            'success': True,
            'portfolio_review': portfolio_review.get('review', {}) if portfolio_review.get('success') else None,
            'stock_recommendations': stock_recommendations.get('recommendations', []) if stock_recommendations.get('success') else [],
            'errors': []
        }
        
        # Add any errors from individual calls
        if not portfolio_review.get('success'):
            result['errors'].append(f"Portfolio review: {portfolio_review.get('error', 'Unknown error')}")
        
        if not stock_recommendations.get('success'):
            result['errors'].append(f"Stock recommendations: {stock_recommendations.get('error', 'Unknown error')}")
        
        # If both failed, mark overall as failed
        if not portfolio_review.get('success') and not stock_recommendations.get('success'):
            result['success'] = False
            result['error'] = '; '.join(result['errors'])
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@portfolio_bp.route('/api/news')
@login_required
def get_portfolio_news_api():
    """Get news for portfolio holdings"""
    try:
        from services.news_service import get_portfolio_news, get_market_news
        
        # Get user's holdings
        holdings = Holding.query.filter_by(user_id=current_user.id).all()
        
        if not holdings:
            # Return market news if no holdings
            news = get_market_news(10)
            return jsonify({
                'success': True,
                'data': news,
                'source': 'market'
            })
        
        # Get unique symbols
        symbols = list(set(h.symbol for h in holdings))
        
        # Fetch news
        news = get_portfolio_news(symbols, limit_per_stock=3, total_limit=15)
        
        return jsonify({
            'success': True,
            'data': news,
            'source': 'portfolio'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@portfolio_bp.route('/export/csv')
@login_required
def export_csv():
    """Export portfolio holdings as CSV"""
    import csv
    from io import StringIO
    from flask import Response
    from services.stock_service import get_current_prices, get_dividend_info, get_day_changes
    
    holdings = Holding.query.filter_by(user_id=current_user.id).all()
    
    if not holdings:
        flash('No holdings to export', 'warning')
        return redirect(url_for('portfolio.index'))
    
    # Consolidate holdings by symbol
    consolidated = {}
    for h in holdings:
        symbol = h.symbol
        if symbol not in consolidated:
            consolidated[symbol] = {
                'symbol': symbol,
                'company_name': h.company_name,
                'quantity': h.quantity,
                'total_cost': h.quantity * h.buy_price,
                'sector': h.sector,
            }
        else:
            consolidated[symbol]['quantity'] += h.quantity
            consolidated[symbol]['total_cost'] += h.quantity * h.buy_price
    
    # Fetch current prices and dividends
    symbols = list(consolidated.keys())
    current_prices = get_current_prices(symbols)
    dividend_info = get_dividend_info(symbols)
    day_changes = get_day_changes(symbols)
    
    # Create CSV in memory
    si = StringIO()
    writer = csv.writer(si)
    
    # Header row
    writer.writerow([
        'Symbol', 'Company', 'Quantity', 'Avg Buy Price', 'Current Price',
        'Invested', 'Current Value', 'Gain/Loss', 'Gain/Loss %', 
        'Day Change', 'Day Change %', 'Annual Dividend', 'Sector'
    ])
    
    # Data rows
    for symbol, data in consolidated.items():
        price_data = current_prices.get(symbol, {'price': 0})
        current_price = price_data['price'] if isinstance(price_data, dict) else price_data
        avg_price = data['total_cost'] / data['quantity'] if data['quantity'] > 0 else 0
        current_value = data['quantity'] * current_price
        gain_loss = current_value - data['total_cost']
        gain_loss_pct = ((current_price - avg_price) / avg_price * 100) if avg_price > 0 else 0
        
        day_change_data = day_changes.get(symbol, {'change': 0, 'change_pct': 0})
        day_change = day_change_data['change'] * data['quantity']
        
        dividend = dividend_info.get(symbol, 0) * data['quantity']
        
        writer.writerow([
            symbol,
            data['company_name'] or '',
            round(data['quantity'], 2),
            round(avg_price, 2),
            round(current_price, 2),
            round(data['total_cost'], 2),
            round(current_value, 2),
            round(gain_loss, 2),
            round(gain_loss_pct, 2),
            round(day_change, 2),
            round(day_change_data['change_pct'], 2),
            round(dividend, 2),
            data['sector'] or 'N/A'
        ])
    
    output = si.getvalue()
    return Response(
        output,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment;filename=portfolio_export.csv'}
    )


@portfolio_bp.route('/export/excel')
@login_required
def export_excel():
    """Export portfolio holdings as Excel"""
    from io import BytesIO
    from flask import send_file
    from services.stock_service import get_current_prices, get_dividend_info, get_day_changes
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('Excel export requires openpyxl. Please install it.', 'error')
        return redirect(url_for('portfolio.index'))
    
    holdings = Holding.query.filter_by(user_id=current_user.id).all()
    
    if not holdings:
        flash('No holdings to export', 'warning')
        return redirect(url_for('portfolio.index'))
    
    # Consolidate holdings by symbol
    consolidated = {}
    for h in holdings:
        symbol = h.symbol
        if symbol not in consolidated:
            consolidated[symbol] = {
                'symbol': symbol,
                'company_name': h.company_name,
                'quantity': h.quantity,
                'total_cost': h.quantity * h.buy_price,
                'sector': h.sector,
            }
        else:
            consolidated[symbol]['quantity'] += h.quantity
            consolidated[symbol]['total_cost'] += h.quantity * h.buy_price
    
    # Fetch current prices and dividends
    symbols = list(consolidated.keys())
    current_prices = get_current_prices(symbols)
    dividend_info = get_dividend_info(symbols)
    day_changes = get_day_changes(symbols)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Holdings"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    currency_style = '#,##0.00'
    percent_style = '0.00%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header row
    headers = [
        'Symbol', 'Company', 'Quantity', 'Avg Buy Price', 'Current Price',
        'Invested', 'Current Value', 'Gain/Loss', 'Gain/Loss %', 
        'Day Change', 'Day Change %', 'Annual Dividend', 'Sector'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    total_invested = 0
    total_current = 0
    total_gain_loss = 0
    
    row = 2
    for symbol, data in consolidated.items():
        price_data = current_prices.get(symbol, {'price': 0})
        current_price = price_data['price'] if isinstance(price_data, dict) else price_data
        avg_price = data['total_cost'] / data['quantity'] if data['quantity'] > 0 else 0
        current_value = data['quantity'] * current_price
        gain_loss = current_value - data['total_cost']
        gain_loss_pct = ((current_price - avg_price) / avg_price) if avg_price > 0 else 0
        
        day_change_data = day_changes.get(symbol, {'change': 0, 'change_pct': 0})
        day_change = day_change_data['change'] * data['quantity']
        
        dividend = dividend_info.get(symbol, 0) * data['quantity']
        
        total_invested += data['total_cost']
        total_current += current_value
        total_gain_loss += gain_loss
        
        row_data = [
            symbol,
            data['company_name'] or '',
            round(data['quantity'], 2),
            round(avg_price, 2),
            round(current_price, 2),
            round(data['total_cost'], 2),
            round(current_value, 2),
            round(gain_loss, 2),
            gain_loss_pct,
            round(day_change, 2),
            day_change_data['change_pct'] / 100,
            round(dividend, 2),
            data['sector'] or 'N/A'
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in [4, 5, 6, 7, 8, 10, 12]:  # Currency columns
                cell.number_format = '₹#,##0.00'
            elif col in [9, 11]:  # Percentage columns
                cell.number_format = '0.00%'
        
        row += 1
    
    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_invested, 2)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(total_current, 2)).font = Font(bold=True)
    ws.cell(row=row, column=8, value=round(total_gain_loss, 2)).font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 20)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='portfolio_export.xlsx'
    )
